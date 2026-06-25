"""Import portfolio holdings from Excel"""
from datetime import datetime, date
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from models import Holding, AssetType, Currency
from crawlers.exchange_rates import guess_currency_from_code, get_rate


def guess_asset_type(code: str) -> str:
    """Guess asset type from security code"""
    code = str(code).strip().upper()

    if code in ('GOOGL', 'NVDA', 'INTC', 'SNDK', 'AMD', 'AAPL', 'MSFT', 'AMZN', 'TSLA'):
        return AssetType.US_STOCK.value
    if code == 'QQQ':
        return AssetType.US_ETF.value

    if code.endswith('.SZ') or code.endswith('.SH'):
        return AssetType.A_SHARE_ETF.value

    if code.endswith('.OF'):
        if code.startswith(('006829', '014856', '006517')):
            return AssetType.BOND.value
        if code.startswith(('008701', '008702', '002611')):
            return AssetType.GOLD.value
        if code.startswith(('019524', '019525', '006479', '015311', '007722')):
            return AssetType.QDII_EQUITY.value
        if code.startswith(('018388', '021142')):
            return AssetType.HK_EQUITY.value
        return AssetType.A_SHARE_EQUITY.value

    return AssetType.CASH.value


def import_excel(filepath: str, db: Session, batch: str | None = None, user_id: int = 1) -> int:
    """
    Import holdings from Excel, deduplicating same code+quantity rows.
    Fund 'amount' is stored as quantity (份额).
    Adds price and calculates amount = quantity × price.

    user_id: 隔离持仓（多用户升级）。只删该 user 的旧持仓。
    """
    if batch is None:
        batch = datetime.utcnow().strftime("%Y%m%d%H%M%S")

    # Delete old holdings for this user（仅删自己的，不影响其他 user）
    db.query(Holding).filter(Holding.user_id == user_id).delete()

    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Dedup map: (code, quantity) → merged record
    dedup: dict[tuple[str, float], dict] = {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        code, name, col2, col3 = row[0], row[1], row[2], row[3]
        if code is None:
            continue

        code = str(code).strip()
        name = str(name).strip() if name else ""
        asset_type = guess_asset_type(code)

        # Determine quantity: all fund amounts ARE quantities (份额)
        if asset_type in (AssetType.US_STOCK.value, AssetType.US_ETF.value):
            # US stocks: column 3 = share count
            quantity = float(col2) if col2 else 0.0
        else:
            # Chinese funds: column 3 = 份额 (quantity, not monetary amount)
            # column 4 is duplicate of column 3
            quantity = float(col3) if col3 else (float(col2) if col2 else 0.0)

        if quantity == 0:
            continue

        key = (code, round(quantity, 4))
        if key in dedup:
            continue  # skip duplicate

        dedup[key] = {
            "user_id": user_id,
            "security_code": code,
            "security_name": name,
            "quantity": quantity,
            "price": None,
            "currency": guess_currency_from_code(code),
            "amount": 0.0,
            "amount_cny": 0.0,
            "asset_type": asset_type,
            "import_batch": batch,
        }

    # Write deduplicated records
    count = 0
    for data in dedup.values():
        h = Holding(**data)
        db.add(h)
        count += 1

    db.commit()
    wb.close()
    return count


def _fetch_fund_nav(fund_code: str) -> float | None:
    """Get latest NAV for a Chinese fund via akshare.
    Tries 单位净值 first, then fallback to 累计净值 for QDII/bond funds."""
    try:
        import akshare as ak
        # Try 单位净值走势 (standard)
        df = ak.fund_open_fund_info_em(symbol=fund_code, indicator="单位净值走势")
        if df is not None and len(df) > 0:
            nav = df.iloc[-1].get("单位净值")
            if nav is not None:
                return float(nav)
    except Exception:
        pass

    try:
        # Fallback: 累计净值走势 (QDII, some bond funds)
        df = ak.fund_open_fund_info_em(symbol=fund_code, indicator="累计净值走势")
        if df is not None and len(df) > 0:
            nav = df.iloc[-1].get("累计净值")
            if nav is not None:
                return float(nav)
    except Exception:
        pass
    return None


def fetch_fund_nav_history(fund_code: str, days: int = 90) -> list[dict]:
    """拉 OF 基金过去 N 天每日净值（真实数据，来自东方财富）。
    返回 [{date: 'YYYY-MM-DD', close: 1.234}, ...]
    优先用「单位净值走势」，单位净值为空的基金（如某些 QDII/债券）用「累计净值走势」。"""
    from datetime import timedelta

    def _try(indicator: str, value_col: str) -> list[dict]:
        # 优先用 DB 映射表（api_code_map.akshare_fund_nav）
        from services.code_map import transform_code
        from database import SessionLocal
        db = SessionLocal()
        try:
            mapped = transform_code(fund_code, "akshare_fund_nav", db)
        finally:
            db.close()
        effective = mapped if mapped else fund_code
        try:
            import akshare as ak
            df = ak.fund_open_fund_info_em(symbol=effective, indicator=indicator)
        except Exception:
            return []
        if df is None or df.empty:
            return []
        out = []
        cutoff = date.today() - timedelta(days=days)
        for _, row in df.iterrows():
            try:
                raw_date = row.get("净值日期")
                if raw_date is None:
                    raw_date = row.get("日期")
                if raw_date is None:
                    continue
                # 统一成 date 对象
                # pandas.Timestamp → date
                # datetime.datetime → date
                # datetime.date → 直接用
                # 字符串 → 解析
                if hasattr(raw_date, "date") and callable(raw_date.date) and not isinstance(raw_date, date):
                    d = raw_date.date()
                elif isinstance(raw_date, date):
                    d = raw_date
                else:
                    d = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
                if d < cutoff:
                    continue
                nav = row.get(value_col)
                if nav is None or str(nav) == "nan":
                    continue
                out.append({"date": d.isoformat(), "close": float(nav)})
            except Exception:
                # 单行解析失败不影响整体
                continue
        return out

    result = _try("单位净值走势", "单位净值")
    if result:
        return result
    return _try("累计净值走势", "累计净值")


def fill_prices(db: Session, user_id: int | None = None):
    """Fetch latest prices for holdings and calculate amount = quantity × price.

    多用户隔离：user_id=None 时处理全部；否则只处理该 user 的 holdings（2026-06-24）。
    价格获取走公共缓冲层（services.price_cache），TTL 15min，多用户共享（2026-06-25）。"""
    from services.price_cache import get_realtime_price
    q = db.query(Holding)
    if user_id is not None:
        q = q.filter(Holding.user_id == user_id)
    holdings = q.all()
    updated = 0

    for h in holdings:
        try:
            price, source, status = get_realtime_price(
                db, h.security_code, h.asset_type, h.currency
            )
            if price and price > 0:
                h.price = round(price, 4)
                h.amount = round(h.quantity * price, 2)
                # Convert to CNY
                rate = get_rate(db, h.currency, 'CNY')
                if rate > 0:
                    h.amount_cny = round(h.amount * rate, 2)
                else:
                    h.amount_cny = h.amount
                if status in ("refreshed", "hit", "nav"):
                    updated += 1
        except Exception:
            pass

    # For holdings still without price or amount_cny: amount = quantity (unit price ≈ 1)
    for h in holdings:
        needs_amount = (h.amount is None or h.amount == 0) and h.quantity > 0
        needs_amount_cny = (h.amount_cny is None or h.amount_cny == 0) and h.quantity > 0
        if needs_amount or needs_amount_cny:
            h.price = h.price or 1.0
            if needs_amount:
                h.amount = round(h.quantity * h.price, 2)
            # 始终重新计算 amount_cny（修复历史 amt_cny=0 的行）
            rate = get_rate(db, h.currency, 'CNY')
            if rate > 0:
                h.amount_cny = round(h.amount * rate, 2)
            else:
                h.amount_cny = h.amount

    db.commit()
    return updated


def get_holdings_summary(db: Session, user_id: int | None = None) -> dict:
    """Get portfolio summary by asset category.

    user_id 隔离：传 None 时返回所有（兼容旧调用；不建议用于受保护端点）。
    """
    from sqlalchemy import func
    q = db.query(Holding)
    if user_id is not None:
        q = q.filter(Holding.user_id == user_id)
    rows = q.with_entities(
        Holding.asset_type,
        func.sum(Holding.amount).label("total")
    ).group_by(Holding.asset_type).all()

    categories = {}
    total = 0.0
    for r in rows:
        v = float(r.total or 0)
        categories[r.asset_type] = v
        total += v

    fund_count = q.filter(
        Holding.asset_type.in_([
            AssetType.A_SHARE_EQUITY.value, AssetType.A_SHARE_ETF.value,
            AssetType.HK_EQUITY.value, AssetType.QDII_EQUITY.value,
            AssetType.US_ETF.value,
        ])
    ).count()

    stock_count = q.filter(
        Holding.asset_type == AssetType.US_STOCK.value
    ).count()

    # 查询最新日 CASH 行（交易形成的现金）。Holding 表不含 CASH，需从 HoldingDailySnapshot 取。
    # 取该用户最新一天的 CASH 行（as_of_date 倒序第一行），用 amount_cny。
    from models import HoldingDailySnapshot
    cash_q = db.query(HoldingDailySnapshot).filter(
        HoldingDailySnapshot.is_cash == True,
    )
    if user_id is not None:
        cash_q = cash_q.filter(HoldingDailySnapshot.user_id == user_id)
    latest_cash = cash_q.order_by(HoldingDailySnapshot.as_of_date.desc()).first()
    cash_cny = float(latest_cash.amount_cny or latest_cash.quantity or 0.0) if latest_cash else 0.0

    return {
        "total_value": round(total, 2),
        "categories": {k: round(v, 2) for k, v in categories.items()},
        "fund_count": fund_count,
        "stock_count": stock_count,
        "cash_cny": round(cash_cny, 2),
    }
